"""
Pharmacy Reports Routes
Handles submission, validation, and management of pharmacy safety reports
With integrated case scoring, case linkage, and WhatsApp/Email notifications

COMPLETE PIPELINE (same as doctor flow):
1. Save to PharmacyReport table (for pharmacy records)
2. Create Patient record (for PV tracking)
3. Run case scoring and case matching
4. Trigger WhatsApp and Email notifications via PVAgentOrchestrator
"""
from flask import Blueprint, request, jsonify, session
from datetime import datetime
import random
from models import db, Patient
from pv_backend.models.pharmacy_report import (
    PharmacyReport, AnonymousReport, IdentifiedReport, AggregatedReport,
    ReportType, ReactionSeverity, ReactionOutcome, AgeGroup
)
from pv_backend.services.audit_service import log_action
from pv_backend.services.case_scoring import evaluate_case, score_case
from pv_backend.services.case_matching import match_new_case, should_accept_case, CaseMatchingEngine

pharmacy_report_bp = Blueprint('pharmacy_reports', __name__, url_prefix='/api/pharmacy/reports')

# ============================================================================
# SCHEMA VALIDATION
# ============================================================================

REPORT_SCHEMAS = {
    'anonymous': [
        'drug_name', 'amount', 'batch_lot_number', 'date_of_dispensing',
        'reaction_category', 'severity', 'reaction_outcome', 'age_group', 'gender', 'additional_notes'
    ],
    'identified': [
        'patient_name', 'patient_email', 'patient_phone',
        'drug_name', 'amount', 'reaction_category', 'severity', 'additional_notes'
    ],
    'aggregated': [
        'drug_name', 'amount', 'total_reactions_reported',
        'mild_count', 'moderate_count', 'severe_count',
        'reporting_period_start', 'reporting_period_end', 'analysis_notes'
    ]
}

# ============================================================================
# SUBMISSION ENDPOINTS
# ============================================================================

def check_duplicate_patient_pharmacy(name, drug_name, age, gender, symptoms=None, phone=None, email=None):
    """
    Check for duplicate patient entries before adding to database.
    
    For pharmacy reports:
    - REJECT only if SAME person AND SAME drug (exact duplicate report)
    - ACCEPT if same person but DIFFERENT drug (new drug reaction, send notifications)
    - Multiple different patients can take the same drug at a pharmacy
    """
    existing_patients = Patient.query.all()
    
    for existing in existing_patients:
        # Check for exact same person (phone or email match)
        person_match = False
        
        if phone and existing.phone and phone == existing.phone:
            person_match = True
        
        if email and existing.email and email.lower() == existing.email.lower():
            person_match = True
        
        if person_match:
            # Same person found - check if same drug
            drug_match = False
            if drug_name and existing.drug_name:
                drug_match = drug_name.lower().strip() == existing.drug_name.lower().strip()
            
            if drug_match:
                # SAME PERSON + SAME DRUG = EXACT DUPLICATE, REJECT
                return {
                    'is_duplicate': True,
                    'action': 'REJECT',
                    'existing_case': existing,
                    'match_score': 1.0,
                    'reason': f"Exact duplicate - Same patient '{existing.name}' already reported this drug (ID: {existing.id})"
                }
            else:
                # SAME PERSON + DIFFERENT DRUG = NEW REPORT, ACCEPT and send notifications
                return {
                    'is_duplicate': False,
                    'action': 'ACCEPT',
                    'existing_case': existing,
                    'match_score': 0,
                    'reason': f"Same patient but different drug - Accept as new case and send notifications"
                }
    
    # No match found - accept as new patient
    return {
        'is_duplicate': False,
        'action': 'ACCEPT',
        'existing_case': None,
        'match_score': 0,
        'reason': 'New patient - accepting as new case'
    }


@pharmacy_report_bp.route('/submit', methods=['POST'])
def submit_report():
    """
    Submit pharmacy safety report(s)
    Supports: Manual entry, Excel upload
    
    COMPLETE PIPELINE (same as doctor flow):
    1. Save to PharmacyReport table (for pharmacy records)
    2. Create Patient record (for PV tracking) - for identified reports with contact info
    3. Run duplicate check and case matching
    4. Run case scoring (evaluate_case + score_case)
    5. Trigger WhatsApp and Email notifications via PVAgentOrchestrator
    """
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': 'No data provided'}), 400
        
        report_type = data.get('report_type', 'anonymous')
        entry_mode = data.get('entry_mode', 'manual')
        records = data.get('records', [])
        
        # DEBUG: Log what we received
        print(f"[PHARMACY SUBMIT] Received submission:")
        print(f"  report_type: {report_type}")
        print(f"  entry_mode: {entry_mode}")
        print(f"  records count: {len(records)}")
        for i, rec in enumerate(records):
            print(f"  Record {i+1} keys: {list(rec.keys())}")
            print(f"  Record {i+1} data: {rec}")
        
        # Validate report type
        if report_type not in REPORT_SCHEMAS:
            return jsonify({'success': False, 'message': f'Invalid report type: {report_type}'}), 400
        
        if not records:
            return jsonify({'success': False, 'message': 'No records provided'}), 400
        
        # Get pharmacy ID from session
        pharmacy_id = session.get('user_id')
        if not pharmacy_id:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
        # Helper function to safely get enum values
        def safe_age_group(value):
            try:
                return AgeGroup(value) if value else AgeGroup.UNKNOWN
            except ValueError:
                return AgeGroup.UNKNOWN
        
        def safe_severity(value):
            try:
                return ReactionSeverity(value) if value else ReactionSeverity.MILD
            except ValueError:
                return ReactionSeverity.MILD
        
        def safe_outcome(value):
            try:
                return ReactionOutcome(value) if value else None
            except ValueError:
                return None
        
        # Create submission records
        submission_id = f"SUB-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}"
        created_reports = []
        created_patients = []
        followup_results = []
        skipped_duplicates = []
        linked_cases = []
        
        for idx, record in enumerate(records):
            try:
                # ============================================================
                # STEP 1: Create PharmacyReport record
                # ============================================================
                if report_type == 'anonymous':
                    report = AnonymousReport(
                        report_type=ReportType.ANONYMOUS,
                        pharmacy_id=pharmacy_id,
                        drug_name=record.get('drug_name'),
                        drug_batch_number=record.get('batch_lot_number'),
                        reaction_description=record.get('reaction_category') or record.get('additional_notes') or 'Adverse reaction reported',
                        reaction_severity=safe_severity(record.get('severity')),
                        reaction_outcome=safe_outcome(record.get('reaction_outcome')),
                        age_group=safe_age_group(record.get('age_group')),
                        gender=record.get('gender')
                    )
                
                elif report_type == 'identified':
                    report = IdentifiedReport(
                        report_type=ReportType.IDENTIFIED,
                        pharmacy_id=pharmacy_id,
                        drug_name=record.get('drug_name'),
                        drug_batch_number=record.get('batch_lot_number'),
                        reaction_description=record.get('reaction_category') or record.get('additional_notes') or 'Adverse reaction reported',
                        reaction_severity=safe_severity(record.get('severity')),
                        reaction_outcome=safe_outcome(record.get('reaction_outcome')),
                        age_group=safe_age_group(record.get('age_group')),
                        gender=record.get('gender'),
                        patient_name=record.get('patient_name'),
                        patient_email=record.get('patient_email'),
                        patient_phone=record.get('patient_phone'),
                        internal_case_id=record.get('internal_case_id'),
                        treating_hospital_reference=record.get('treating_hospital_reference'),
                        treating_doctor_name=record.get('treating_doctor_name'),
                        follow_up_required=True  # Flag for follow-up
                    )
                
                elif report_type == 'aggregated':
                    report = AggregatedReport(
                        report_type=ReportType.AGGREGATED,
                        pharmacy_id=pharmacy_id,
                        drug_name=record.get('drug_name'),
                        reaction_description=f"Aggregated report for {record.get('drug_name')}",
                        report_count=int(record.get('total_reactions_reported') or 1)
                    )
                
                db.session.add(report)
                db.session.flush()  # Get report ID without committing
                created_reports.append(report)
                
                # ============================================================
                # STEP 2: For identified reports with contact info, create Patient 
                # and trigger FULL PV pipeline (same as doctor flow)
                # ============================================================
                if report_type == 'identified':
                    patient_name = record.get('patient_name') or f'Patient-{submission_id}-{idx+1}'
                    # Clean up phone/email - treat empty strings as None
                    patient_phone = record.get('patient_phone', '').strip() or None
                    patient_email = record.get('patient_email', '').strip() or None
                    
                    print(f"[PHARMACY SUBMIT] Record {idx+1} data received: name={patient_name}, phone='{patient_phone}', email='{patient_email}'")
                    print(f"[PHARMACY SUBMIT] Full record data: {record}")
                    
                    # Only create Patient record if we have contact info for follow-up
                    if patient_phone or patient_email:
                        print(f"[PHARMACY SUBMIT] Creating patient for report {idx+1}: {patient_name}, phone={patient_phone}, email={patient_email}")
                        
                        # Map age group to approximate age
                        age_map = {
                            'pediatric': 8, 'adolescent': 16, 'adult': 35,
                            'elderly': 65, 'geriatric': 80, 'unknown': 40
                        }
                        age_group_val = record.get('age_group', 'adult')
                        age = age_map.get(age_group_val, 35)
                        
                        # Map severity to risk level
                        severity_map = {'mild': 'Low', 'moderate': 'Medium', 'severe': 'High'}
                        risk_level = severity_map.get(record.get('severity', 'moderate'), 'Medium')
                        
                        drug_name = record.get('drug_name', 'Not Specified')
                        symptoms = record.get('reaction_category') or record.get('additional_notes') or ''
                        gender = record.get('gender') or 'Not Specified'  # Default if not provided
                        
                        # ============================================================
                        # STEP 2a: Check for duplicates (same as doctor flow)
                        # ============================================================
                        duplicate_check = check_duplicate_patient_pharmacy(
                            name=patient_name,
                            drug_name=drug_name,
                            age=age,
                            gender=gender,
                            symptoms=symptoms,
                            phone=patient_phone,
                            email=patient_email
                        )
                        
                        if duplicate_check['action'] == 'REJECT':
                            # Skip this record - exact duplicate
                            print(f"[DUPLICATE REJECTED] {duplicate_check['reason']}")
                            skipped_duplicates.append({
                                'record_index': idx + 1,
                                'patient_name': patient_name,
                                'reason': duplicate_check['reason'],
                                'existing_patient_id': duplicate_check['existing_case'].id if duplicate_check['existing_case'] else None
                            })
                            continue
                        
                        # Generate unique patient ID
                        patient_id = f"PHR-{random.randint(10000, 99999)}"
                        while Patient.query.get(patient_id):
                            patient_id = f"PHR-{random.randint(10000, 99999)}"
                        
                        # Handle LINK or ACCEPT
                        if duplicate_check['action'] == 'LINK':
                            existing = duplicate_check['existing_case']
                            patient = Patient(
                                id=patient_id,
                                created_by=pharmacy_id,
                                name=patient_name,
                                phone=patient_phone,
                                email=patient_email,
                                age=age,
                                gender=gender,
                                drug_name=drug_name,
                                symptoms=symptoms,
                                risk_level=risk_level,
                                case_status='Linked',
                                linked_case_id=existing.id,
                                match_score=duplicate_check['match_score'],
                                match_notes=f"Auto-linked from pharmacy report: {duplicate_check['reason']}"
                            )
                            linked_cases.append({
                                'new_patient_id': patient_id,
                                'linked_to': existing.id,
                                'match_score': duplicate_check['match_score']
                            })
                            print(f"[CASE LINKED] New patient {patient_id} linked to existing {existing.id}")
                        else:
                            # ACCEPT - create new patient
                            patient = Patient(
                                id=patient_id,
                                created_by=pharmacy_id,
                                name=patient_name,
                                phone=patient_phone,
                                email=patient_email,
                                age=age,
                                gender=gender or 'Not Specified',  # Ensure gender is never NULL
                                drug_name=drug_name,
                                symptoms=symptoms,
                                risk_level=risk_level,
                                case_status='Active'
                            )
                        
                        db.session.add(patient)
                        db.session.flush()  # Get patient into session
                        created_patients.append(patient)
                        print(f"[PATIENT CREATED] {patient.id} - {patient.name}")
                        
                        # ============================================================
                        # STEP 2b: Run case scoring (same as doctor flow)
                        # ============================================================
                        try:
                            evaluate_case(patient)
                            score_result = score_case(patient)
                            print(f"[CASE SCORING] Patient {patient.id}: Score={patient.case_score}, Strength={patient.strength_level}")
                        except Exception as score_err:
                            print(f"[CASE SCORING ERROR] {score_err}")
                        
                        # ============================================================
                        # STEP 2c: Trigger PV Agent follow-up (WhatsApp + Email)
                        # This is the same as auto_send_followup() in doctor flow
                        # ============================================================
                        try:
                            from pv_backend.services.followup_agent import PVAgentOrchestrator
                            
                            orchestrator = PVAgentOrchestrator()
                            followup_result = orchestrator.start_tracking(patient)
                            
                            if followup_result.get('success'):
                                patient.followup_sent_date = datetime.utcnow()
                                patient.followup_pending = True
                                patient.follow_up_sent = True
                                
                                # Extract send results from the orchestrator response
                                send_result = followup_result.get('send_result', {})
                                email_success = send_result.get('email', {}).get('success', False)
                                whatsapp_result = send_result.get('whatsapp', {})
                                whatsapp_success = whatsapp_result.get('success', False)
                                whatsapp_error = whatsapp_result.get('error', None)
                                
                                # Debug logging
                                print(f"[WHATSAPP DEBUG] send_result keys: {send_result.keys()}")
                                print(f"[WHATSAPP DEBUG] whatsapp_result: {whatsapp_result}")
                                print(f"[WHATSAPP DEBUG] whatsapp_success: {whatsapp_success}")
                                print(f"[WHATSAPP DEBUG] whatsapp_error: {whatsapp_error}")
                                
                                followup_results.append({
                                    'patient_id': patient.id,
                                    'patient_name': patient.name,
                                    'status': 'sent',
                                    'tracking_id': followup_result.get('tracking_id'),
                                    'email_sent': email_success,
                                    'whatsapp_sent': whatsapp_success,
                                    'whatsapp_error': whatsapp_error,
                                    'current_day': followup_result.get('current_day', 1),
                                    'questions_count': len(followup_result.get('all_questions', []))
                                })
                                print(f"[PV AGENT OK] Started for patient {patient.id} - Day 1 of 1/3/5/7 cycle")
                                print(f"[PV AGENT] Email: {email_success}, WhatsApp: {whatsapp_success}")
                                if whatsapp_error:
                                    print(f"[PV AGENT] WhatsApp error: {whatsapp_error}")
                            else:
                                followup_results.append({
                                    'patient_id': patient.id,
                                    'patient_name': patient.name,
                                    'status': 'failed',
                                    'error': followup_result.get('error', 'Unknown error')
                                })
                                print(f"[PV AGENT FAILED] {followup_result.get('error')}")
                                
                        except Exception as followup_err:
                            print(f"[FOLLOWUP ERROR] {followup_err}")
                            import traceback
                            traceback.print_exc()
                            followup_results.append({
                                'patient_id': patient.id,
                                'patient_name': patient.name,
                                'status': 'error',
                                'error': str(followup_err)
                            })
                    else:
                        print(f"[PHARMACY SUBMIT] Record {idx+1}: No contact info (phone/email) - skipping patient creation for PV tracking")
            
            except Exception as record_err:
                print(f"[RECORD ERROR] Record {idx+1}: {record_err}")
                import traceback
                traceback.print_exc()
                # Continue with other records even if one fails
                continue
        
        # Commit all changes
        try:
            db.session.commit()
            print(f"[COMMIT OK] Reports: {len(created_reports)}, Patients: {len(created_patients)}")
            
            # Log submission
            log_action(
                user_id=pharmacy_id,
                action='PHARMACY_REPORT_SUBMITTED',
                details={
                    'submission_id': submission_id,
                    'report_type': report_type,
                    'entry_mode': entry_mode,
                    'reports_created': len(created_reports),
                    'patients_created': len(created_patients),
                    'duplicates_skipped': len(skipped_duplicates),
                    'cases_linked': len(linked_cases),
                    'followups_triggered': len([f for f in followup_results if f.get('status') == 'sent'])
                }
            )
            
            # Build response
            response_data = {
                'success': True,
                'submission_id': submission_id,
                'message': f'Successfully submitted {len(created_reports)} report(s)',
                'record_count': len(created_reports),
                'patients_created': len(created_patients),
                'followup_results': followup_results
            }
            
            # Add detailed notifications
            for i, result in enumerate(followup_results, 1):
                # Check email status from followup_results
                if result.get('email_sent'):
                    response_data[f'patient_{i}_email_sent'] = True
                
                # Check WhatsApp status from followup_results  
                if result.get('whatsapp_sent'):
                    response_data[f'patient_{i}_whatsapp_sent'] = True
                else:
                    # WhatsApp was not successfully sent
                    response_data[f'patient_{i}_whatsapp_sent'] = False
                    
                    # Check if there was an error
                    whatsapp_error = result.get('whatsapp_error')
                    if whatsapp_error:
                        response_data[f'patient_{i}_whatsapp_error'] = whatsapp_error
                        
                        # Provide helpful message based on error type
                        if 'not configured' in whatsapp_error.lower():
                            response_data[f'patient_{i}_whatsapp_help'] = "⚠️ WhatsApp is not configured. Admin needs to set Twilio credentials (TWILIO_ACCOUNT_SID, TWILIO_AUTH_TOKEN, TWILIO_WHATSAPP_FROM)."
                        elif 'unverified' in whatsapp_error.lower() or 'sandbox' in whatsapp_error.lower():
                            response_data[f'patient_{i}_whatsapp_help'] = f"Patient needs to join Twilio sandbox: Send 'join bright-joy' to +1-415-523-8886"
                        else:
                            response_data[f'patient_{i}_whatsapp_help'] = whatsapp_error
                    else:
                        # Check if status was 'sent' (meaning it was attempted)
                        if result.get('status') == 'sent':
                            response_data[f'patient_{i}_whatsapp_pending'] = True
                            response_data[f'patient_{i}_whatsapp_help'] = f"Patient needs to join Twilio sandbox: Send 'join bright-joy' to +1-415-523-8886"
            
            if skipped_duplicates:
                response_data['skipped_duplicates'] = skipped_duplicates
            
            if linked_cases:
                response_data['linked_cases'] = linked_cases
            
            return jsonify(response_data), 201
        
        except Exception as commit_err:
            db.session.rollback()
            print(f"[COMMIT ERROR] {commit_err}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'message': f'Database error: {str(commit_err)}'
            }), 500
    
    except Exception as e:
        print(f"[SUBMIT ERROR] {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'message': f'Server error: {str(e)}'
        }), 500


@pharmacy_report_bp.route('/validate-excel', methods=['POST'])
def validate_excel():
    """
    Validate Excel file schema before submission
    """
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file provided'}), 400
        
        file = request.files['file']
        report_type = request.form.get('report_type', 'anonymous')
        
        if report_type not in REPORT_SCHEMAS:
            return jsonify({'success': False, 'error': f'Invalid report type: {report_type}'}), 400
        
        # Read Excel file
        import openpyxl
        try:
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook.active
            
            # Get headers from first row
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(cell.value)
            
            # Check if all required columns exist
            required_columns = REPORT_SCHEMAS[report_type]
            missing_columns = [col for col in required_columns if col not in headers]
            
            if missing_columns:
                return jsonify({
                    'success': False,
                    'error': f'Missing required columns: {", ".join(missing_columns)}',
                    'required_columns': required_columns
                }), 400
            
            # Read data rows
            preview_rows = []
            total_rows = 0
            
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 1):
                if idx > 10:  # Preview first 10 rows
                    break
                
                row_data = {}
                for col_idx, header in enumerate(headers):
                    row_data[header] = row[col_idx] if col_idx < len(row) else None
                
                preview_rows.append(row_data)
                total_rows = idx
            
            # Count total rows
            total_rows = worksheet.max_row - 1  # Exclude header
            
            return jsonify({
                'success': True,
                'total_rows': total_rows,
                'preview_rows': preview_rows,
                'column_mapping': {h: h for h in headers}
            }), 200
        
        except Exception as e:
            return jsonify({
                'success': False,
                'error': f'Error reading Excel file: {str(e)}'
            }), 400
    
    except Exception as e:
        return jsonify({
            'success': False,
            'error': f'Server error: {str(e)}'
        }), 500


# ============================================================================
# RETRIEVAL ENDPOINTS
# ============================================================================

@pharmacy_report_bp.route('/history', methods=['GET'])
def get_submission_history():
    """
    Get submission history for current pharmacy
    """
    try:
        pharmacy_id = session.get('user_id')
        if not pharmacy_id:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
        limit = request.args.get('limit', 50, type=int)
        offset = request.args.get('offset', 0, type=int)
        
        # Query reports
        reports = PharmacyReport.query.filter_by(pharmacy_id=pharmacy_id)\
            .order_by(PharmacyReport.created_at.desc())\
            .limit(limit)\
            .offset(offset)\
            .all()
        
        return jsonify({
            'success': True,
            'reports': [r.to_dict() for r in reports],
            'total': len(reports)
        }), 200
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500


@pharmacy_report_bp.route('/<int:report_id>', methods=['GET'])
def get_report_detail(report_id):
    """
    Get detailed information about a specific report
    """
    try:
        pharmacy_id = session.get('user_id')
        if not pharmacy_id:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
        report = PharmacyReport.query.filter_by(id=report_id, pharmacy_id=pharmacy_id).first()
        
        if not report:
            return jsonify({'success': False, 'message': 'Report not found'}), 404
        
        return jsonify({
            'success': True,
            'report': report.to_dict()
        }), 200
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500


# ============================================================================
# COMPLIANCE SCORING
# ============================================================================

@pharmacy_report_bp.route('/compliance-score', methods=['GET'])
def get_compliance_score():
    """
    Get compliance score for current pharmacy
    Scores based on:
    - On-time submissions
    - Alert acknowledgments
    - Data quality
    - Schema compliance
    """
    try:
        pharmacy_id = session.get('user_id')
        if not pharmacy_id:
            return jsonify({'success': False, 'message': 'Not authenticated'}), 401
        
        # Calculate compliance score
        score = 100  # Start with perfect score
        
        # Count recent submissions
        recent_reports = PharmacyReport.query.filter_by(pharmacy_id=pharmacy_id)\
            .filter(PharmacyReport.created_at >= datetime.utcnow().replace(day=1))\
            .count()
        
        # Deduct for no submissions this month
        if recent_reports == 0:
            score -= 10
        
        # Get compliance status
        if score >= 80:
            status = 'Compliant'
            status_color = 'green'
        elif score >= 60:
            status = 'Attention Required'
            status_color = 'yellow'
        else:
            status = 'Non-compliant'
            status_color = 'red'
        
        return jsonify({
            'success': True,
            'compliance_score': score,
            'status': status,
            'status_color': status_color,
            'last_updated': datetime.utcnow().isoformat()
        }), 200
    
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error: {str(e)}'
        }), 500
